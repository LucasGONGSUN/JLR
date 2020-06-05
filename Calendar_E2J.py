import openpyxl
import json

# Settings
WB = openpyxl.load_workbook('Calendar.xlsx')
WS = WB['Jap']
rows = WS.max_row
cols = WS.max_column
NumberList = []
NoteList = []
CheckList = {}


# === Create a learning list, save as a dict === #
def Createlist():
    # Read No. from calendar
    for row in WS.iter_rows(min_row=2, max_row=rows, max_col=1):
        for cell in row:
            NumberList.append(cell.value)

    # Read notelist from calendar
    for row in WS.iter_rows(min_row=2, max_row=rows, min_col=2, max_col=cols):
        new_row = []
        for cell in row:
            new_row.append(cell.value)
        NoteList.append(new_row)

    # Combine Number and notelist
    for i in range(184):
        CheckList[str(NumberList[i])] = NoteList[i]


# === Save dict as json files === #
def Dict2Json():
    FileJsonName = 'Calendar.json'
    with open(FileJsonName, 'w') as file_obj:
        json.dump(CheckList, file_obj, indent=2)


# === Save calendar to Json === #
def Save2Json():
    Createlist()
    Dict2Json()
    print('\n>>> Calendar is saved to Json file. <<<')
