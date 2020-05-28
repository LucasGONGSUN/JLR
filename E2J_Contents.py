import openpyxl
import json

# Setting
WB = openpyxl.load_workbook('contents.xlsx')
WS = WB['content']
DateList = []
NoteList = []
ContentList = {}


# === Create a learning list, save as a dict === #
def Createlist():
    global ContentList
    # Read date from calendar
    for row in WS.iter_rows(min_row=1, max_row=151, max_col=1):
        for cell in row:
            DateList.append(cell.value)
    # Read notelist from calendar
    for row in WS.iter_rows(min_row=1, max_row=151, min_col=2, max_col=16):
        new_row = []
        for cell in row:
            new_row.append(cell.value)
        NoteList.append(new_row)
    # Combine date and notelist
    for i in range(151):
        ContentList[str(DateList[i])] = NoteList[i]


# === Save dict as json files === #
def Dict2Json():
    FileJsonName = 'contents.json'
    with open(FileJsonName, 'w') as file_obj:
        json.dump(ContentList, file_obj)

Createlist()
Dict2Json()
print('Mission Completed!')
