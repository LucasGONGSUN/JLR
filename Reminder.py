import openpyxl, datetime, smtplib, schedule
from email.mime.text import MIMEText
from email.header import Header

#设置操作对象
wb = openpyxl.load_workbook('calendar.xlsx')
ws = wb['Jap']
datelist = []
notelist = []
checklist = {}

# === 构筑列表 === #
def getquest():
    global todayquest
    #读取日期
    for row in ws.iter_rows(min_row=1, max_col=1, max_row=185):
        for cell in row:
            datelist.append(cell.value)
    #读取笔记序列
    for row in ws.iter_rows(min_row=1, max_row=185, min_col=2, max_col=8):
        new_row = []
        for cell in row:
            new_row.append(cell.value)
        notelist.append(new_row)
    #组合日期与笔记序列
    for i in range(185):
        checklist[datelist[i]] = notelist[i]

    # === 读取今日列表 === #
    a = datetime.datetime.today()
    today = datetime.datetime(a.year, a.month, a.day, 0,0,0)
    todaylist = checklist[today]
    todaylist_str = "；".join(todaylist)
    todayquest = '今天的学习任务是：' + todaylist_str

# === 生成邮件 === #
def sendmail():
    #设置邮箱信息
    from_addr = '37870979@qq.com'
    password = 'fcapyvxmjdbzbgji'
    to_addr = 'lucasgongsun@163.com'
    smtp_server = 'smtp.qq.com'
    #设置正文信息
    msg = MIMEText(todayquest, 'plain', 'utf-8')
    #设置邮件头信息
    msg['From'] = Header(from_addr)
    msg['To'] = Header(to_addr)
    msg['Subject'] = Header('今日学习任务')
    #开启发信服务
    server = smtplib.SMTP_SSL(host=smtp_server)
    server.connect(smtp_server,465)
    server.login(from_addr, password)
    server.sendmail(from_addr, to_addr, msg.as_string())
    server.quit()

# === 定时发送 === #
def job():
    getquest()
    sendmail()

schedule.every().day.at("8:00").do(job)

